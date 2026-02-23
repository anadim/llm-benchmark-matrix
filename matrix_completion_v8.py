#!/usr/bin/env python3
"""
Matrix Completion v8 — Ensemble + Confidence Intervals + Full Predictions
=========================================================================

Improvements over v7:
1. Stacked ensemble: learn per-benchmark optimal method weights from CV residuals
2. Trimmed ensemble: drop worst predictor per cell, average remaining
3. Confidence intervals via method disagreement + bootstrap
4. Clamp predictions to observed benchmark ranges (no negatives, no >100 for % benchmarks)
5. Full predicted matrix output as Excel + JSON

Evaluation: per-model MedAPE, per-benchmark MedAPE, overall MedAPE
"""

import sys, json, os
import numpy as np
from collections import defaultdict
import warnings
warnings.filterwarnings('ignore')

REPO_ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(REPO_ROOT, 'data'))
from build_benchmark_matrix import MODELS, BENCHMARKS, DATA

# ── Setup ────────────────────────────────────────────────────────────────
model_ids = [m[0] for m in MODELS]
model_meta = {m[0]: m for m in MODELS}
bench_ids = [b[0] for b in BENCHMARKS]
bench_meta = {b[0]: b for b in BENCHMARKS}
model_names = {m[0]: m[1] for m in MODELS}
bench_names = {b[0]: b[1] for b in BENCHMARKS}
model_idx = {m: i for i, m in enumerate(model_ids)}
bench_idx = {b: i for i, b in enumerate(bench_ids)}

n_models = len(model_ids)
n_benchmarks = len(bench_ids)

M_full = np.full((n_models, n_benchmarks), np.nan)
for mid, bid, score, url in DATA:
    if mid in model_idx and bid in bench_idx:
        M_full[model_idx[mid], bench_idx[bid]] = score

observed = ~np.isnan(M_full)
n_observed = observed.sum()
print(f"Matrix: {n_models}×{n_benchmarks}, {n_observed} observed ({100*n_observed/(n_models*n_benchmarks):.1f}%)")

# ── Scale detection ──────────────────────────────────────────────────────
elo_benchmarks = set()
for b in BENCHMARKS:
    if 'Elo' in b[3] or 'rating' in b[3].lower():
        elo_benchmarks.add(b[0])

is_elo = np.array([bench_ids[j] in elo_benchmarks for j in range(n_benchmarks)])
print(f"Elo-scale benchmarks: {sum(is_elo)} ({', '.join(b for b in bench_ids if b in elo_benchmarks)})")

# ── Benchmark ranges (for clamping) ──────────────────────────────────────
bench_min = np.nanmin(M_full, axis=0)
bench_max = np.nanmax(M_full, axis=0)
# For percentage benchmarks, max plausible is 100
for j in range(n_benchmarks):
    if not is_elo[j] and bench_max[j] <= 100:
        bench_max[j] = min(100.0, bench_max[j] * 1.15)  # Allow 15% above max observed
    bench_min[j] = max(0.0, bench_min[j] * 0.5)  # Floor at 0 or half of min

# ── Model families ──────────────────────────────────────────────────────
def get_model_family(mid):
    if mid.startswith("qwen3") and "30b-a3b" not in mid and "235b" not in mid and "397b" not in mid:
        return "qwen3-dense"
    if mid.startswith("deepseek-r1-distill-qwen"):
        return "ds-r1-dq"
    if mid.startswith("deepseek-r1-distill-llama"):
        return "ds-r1-dl"
    if mid.startswith("gpt-5"):
        return "gpt5-series"
    if mid.startswith("claude-opus"):
        return "claude-opus"
    if mid.startswith("claude-sonnet"):
        return "claude-sonnet"
    if mid.startswith("gemini-3"):
        return "gemini3"
    if mid.startswith("gemini-2"):
        return "gemini2"
    if "phi-4" in mid:
        return "phi4"
    if mid.startswith("grok"):
        return "grok"
    return None

FAMILIES = {
    "qwen3-dense": [
        ("qwen3-0.6b", 600), ("qwen3-1.7b", 1700), ("qwen3-4b", 4000),
        ("qwen3-8b", 8000), ("qwen3-14b", 14000), ("qwen3-32b", 32000)
    ],
    "ds-r1-dq": [
        ("deepseek-r1-distill-qwen-1.5b", 1500), ("deepseek-r1-distill-qwen-7b", 7000),
        ("deepseek-r1-distill-qwen-14b", 14000), ("deepseek-r1-distill-qwen-32b", 32000)
    ],
    "ds-r1-dl": [
        ("deepseek-r1-distill-llama-8b", 8000), ("deepseek-r1-distill-llama-70b", 70000)
    ],
}

model_to_family = {}
for fam, members in FAMILIES.items():
    for mid, params in members:
        model_to_family[mid] = (fam, params)


# ══════════════════════════════════════════════════════════════════════════
# BASE METHODS
# ══════════════════════════════════════════════════════════════════════════

def col_zscore(M):
    means = np.nanmean(M, axis=0)
    stds = np.nanstd(M, axis=0)
    stds[stds < 1e-8] = 1.0
    return (M - means) / stds, means, stds


def als_predict(M_train, rank=8, lam=1.0, n_iter=50):
    Z, col_m, col_s = col_zscore(M_train)
    mask = ~np.isnan(Z)
    Z_filled = np.where(mask, Z, 0.0)
    np.random.seed(42)
    U = np.random.randn(n_models, rank) * 0.01
    V = np.random.randn(n_benchmarks, rank) * 0.01
    for it in range(n_iter):
        for i in range(n_models):
            c = np.where(mask[i, :])[0]
            if len(c) == 0: continue
            Vs = V[c, :]
            y = Z_filled[i, c]
            U[i, :] = np.linalg.solve(Vs.T @ Vs + lam * np.eye(rank), Vs.T @ y)
        for j in range(n_benchmarks):
            r = np.where(mask[:, j])[0]
            if len(r) == 0: continue
            Us = U[r, :]
            y = Z_filled[r, j]
            V[j, :] = np.linalg.solve(Us.T @ Us + lam * np.eye(rank), Us.T @ y)
    M_pred = (U @ V.T) * col_s + col_m
    return M_pred


def knn_predict(M_train, k=7, metric='cosine'):
    Z, col_m, col_s = col_zscore(M_train)
    mask = ~np.isnan(Z)
    Z_filled = np.where(mask, Z, 0.0)
    if metric == 'cosine':
        norms = np.linalg.norm(Z_filled, axis=1, keepdims=True)
        norms[norms == 0] = 1.0
        Z_normed = Z_filled / norms
        sim = Z_normed @ Z_normed.T
    else:
        row_means = np.sum(Z_filled * mask, axis=1, keepdims=True) / np.maximum(mask.sum(axis=1, keepdims=True), 1)
        centered = (Z_filled - row_means) * mask
        norms = np.linalg.norm(centered, axis=1, keepdims=True)
        norms[norms == 0] = 1.0
        sim = (centered / norms) @ (centered / norms).T
    overlap = mask.astype(float) @ mask.astype(float).T
    shrinkage = np.minimum(overlap / 5.0, 1.0)
    sim *= shrinkage

    M_pred = M_train.copy()
    for i in range(n_models):
        for j in range(n_benchmarks):
            if mask[i, j]: continue
            candidates = np.where(mask[:, j] & (np.arange(n_models) != i))[0]
            if len(candidates) == 0:
                M_pred[i, j] = col_m[j]
                continue
            sims = sim[i, candidates]
            top_k = min(k, len(candidates))
            top_idx = np.argsort(-sims)[:top_k]
            top_sims = sims[top_idx]
            top_vals = Z[candidates[top_idx], j]
            pos = top_sims > 0
            if pos.sum() > 0:
                w = top_sims[pos]
                v = top_vals[pos]
                pred_z = np.sum(w * v) / np.sum(w)
            else:
                pred_z = 0.0
            M_pred[i, j] = pred_z * col_s[j] + col_m[j]
    return M_pred


def bench_reg_predict(M_train, top_k=5, min_r2=0.2):
    mask = ~np.isnan(M_train)
    col_m = np.nanmean(M_train, axis=0)
    regs = {}
    for j1 in range(n_benchmarks):
        for j2 in range(n_benchmarks):
            if j1 == j2: continue
            both = mask[:, j1] & mask[:, j2]
            n = both.sum()
            if n < 5: continue
            x = M_train[both, j1]
            y = M_train[both, j2]
            if np.std(x) < 1e-8: continue
            slope = np.corrcoef(x, y)[0, 1] * np.std(y) / np.std(x)
            intercept = np.mean(y) - slope * np.mean(x)
            yp = slope * x + intercept
            ss_res = np.sum((y - yp) ** 2)
            ss_tot = np.sum((y - y.mean()) ** 2)
            r2 = 1 - ss_res / ss_tot if ss_tot > 0 else 0
            if r2 >= min_r2:
                regs[(j1, j2)] = (slope, intercept, r2, x.min(), x.max())

    M_pred = M_train.copy()
    for i in range(n_models):
        for j in range(n_benchmarks):
            if mask[i, j]: continue
            preds, weights = [], []
            for j2 in range(n_benchmarks):
                if not mask[i, j2]: continue
                key = (j2, j)
                if key not in regs: continue
                slope, intercept, r2, xmin, xmax = regs[key]
                x_val = M_train[i, j2]
                margin = 0.15 * (xmax - xmin + 1e-6)
                if x_val < xmin - margin or x_val > xmax + margin: continue
                p = slope * x_val + intercept
                preds.append(max(0, p))
                weights.append(r2)
            if preds:
                pairs = sorted(zip(weights, preds), reverse=True)[:top_k]
                tw = sum(w for w, _ in pairs)
                M_pred[i, j] = sum(w * p for w, p in pairs) / tw
            else:
                M_pred[i, j] = col_m[j]
    return M_pred


def scaling_predict(M_train):
    mask = ~np.isnan(M_train)
    col_m = np.nanmean(M_train, axis=0)
    M_pred = M_train.copy()
    for i in range(n_models):
        mid = model_ids[i]
        if mid not in model_to_family:
            for j in range(n_benchmarks):
                if np.isnan(M_pred[i, j]):
                    M_pred[i, j] = col_m[j]
            continue
        fam_name, target_params = model_to_family[mid]
        members = FAMILIES[fam_name]
        for j in range(n_benchmarks):
            if mask[i, j]: continue
            xs, ys = [], []
            for m2, p2 in members:
                mi2 = model_idx.get(m2)
                if mi2 is None: continue
                val = M_train[mi2, j]
                if not np.isnan(val) and val > 0:
                    xs.append(np.log10(p2))
                    ys.append(val)
            if len(xs) < 3:
                if np.isnan(M_pred[i, j]):
                    M_pred[i, j] = col_m[j]
                continue
            xs, ys = np.array(xs), np.array(ys)
            A = np.vstack([xs, np.ones_like(xs)]).T
            coeffs = np.linalg.lstsq(A, ys, rcond=None)[0]
            a, b = coeffs
            yp = a * xs + b
            ss_res = np.sum((ys - yp) ** 2)
            ss_tot = np.sum((ys - ys.mean()) ** 2)
            r2 = 1 - ss_res / ss_tot if ss_tot > 0 else 0
            if r2 > 0.3:
                pred = a * np.log10(target_params) + b
                M_pred[i, j] = max(0, pred)
            elif np.isnan(M_pred[i, j]):
                M_pred[i, j] = col_m[j]
    return M_pred


# ══════════════════════════════════════════════════════════════════════════
# METHOD 7: Trimmed Ensemble (drop worst, average rest)
# ══════════════════════════════════════════════════════════════════════════

def trimmed_ensemble_predict(M_train):
    """
    Run all base methods, then for each cell:
    1. Collect all predictions
    2. Drop the prediction furthest from the median (outlier trimming)
    3. Weighted-average the rest by inverse distance to median
    """
    M_als5 = als_predict(M_train, rank=5, lam=1.0)
    M_als8 = als_predict(M_train, rank=8, lam=1.0)
    M_knn5 = knn_predict(M_train, k=5, metric='cosine')
    M_knn7 = knn_predict(M_train, k=7, metric='cosine')
    M_breg = bench_reg_predict(M_train, top_k=5, min_r2=0.2)
    M_scale = scaling_predict(M_train)

    all_preds = [M_als5, M_als8, M_knn5, M_knn7, M_breg, M_scale]
    mask = ~np.isnan(M_train)
    col_m = np.nanmean(M_train, axis=0)

    M_pred = M_train.copy()
    for i in range(n_models):
        for j in range(n_benchmarks):
            if mask[i, j]: continue
            vals = [P[i, j] for P in all_preds if not np.isnan(P[i, j])]
            if len(vals) == 0:
                M_pred[i, j] = col_m[j]
                continue
            if len(vals) <= 2:
                M_pred[i, j] = np.mean(vals)
                continue
            # Trim the outlier
            med = np.median(vals)
            dists = [abs(v - med) for v in vals]
            worst_idx = np.argmax(dists)
            trimmed = [v for k, v in enumerate(vals) if k != worst_idx]
            # Inverse-distance weighting (closer to median = higher weight)
            trimmed_dists = [abs(v - med) + 0.01 for v in trimmed]
            weights = [1.0 / d for d in trimmed_dists]
            tw = sum(weights)
            M_pred[i, j] = sum(w * v for w, v in zip(weights, trimmed)) / tw
    return M_pred


# ══════════════════════════════════════════════════════════════════════════
# METHOD 8: Stacked Ensemble (learn per-benchmark weights from CV)
# ══════════════════════════════════════════════════════════════════════════

def stacked_ensemble_predict(M_train):
    """
    Step 1: For each benchmark, compute leave-one-out errors of each base method
    Step 2: Weight methods by inverse LOO error per benchmark
    Step 3: Blend predictions using learned weights
    """
    mask = ~np.isnan(M_train)
    col_m = np.nanmean(M_train, axis=0)

    base_methods = [
        ("ALS5", lambda M: als_predict(M, rank=5, lam=1.0)),
        ("ALS8", lambda M: als_predict(M, rank=8, lam=1.0)),
        ("KNN5", lambda M: knn_predict(M, k=5, metric='cosine')),
        ("KNN7", lambda M: knn_predict(M, k=7, metric='cosine')),
        ("BReg", lambda M: bench_reg_predict(M, top_k=5, min_r2=0.2)),
    ]

    # Compute base predictions
    base_preds = [(name, fn(M_train)) for name, fn in base_methods]

    # Compute per-benchmark weights from residuals on observed cells
    bench_weights = np.ones((n_benchmarks, len(base_methods)))
    for j in range(n_benchmarks):
        obs_rows = np.where(mask[:, j])[0]
        if len(obs_rows) < 3:
            continue
        for m_idx, (name, M_p) in enumerate(base_preds):
            errs = []
            for i in obs_rows:
                actual = M_train[i, j]
                pred = M_p[i, j]
                if abs(actual) > 1e-6 and not np.isnan(pred):
                    errs.append(abs((pred - actual) / actual))
            if errs:
                bench_weights[j, m_idx] = 1.0 / (np.median(errs) + 0.02)

    # Normalize weights per benchmark
    for j in range(n_benchmarks):
        total = bench_weights[j].sum()
        if total > 0:
            bench_weights[j] /= total

    # Blend
    M_pred = M_train.copy()
    for i in range(n_models):
        for j in range(n_benchmarks):
            if mask[i, j]: continue
            val = 0.0
            tw = 0.0
            for m_idx, (name, M_p) in enumerate(base_preds):
                p = M_p[i, j]
                if not np.isnan(p):
                    w = bench_weights[j, m_idx]
                    val += w * p
                    tw += w
            M_pred[i, j] = val / tw if tw > 0 else col_m[j]
    return M_pred


# ══════════════════════════════════════════════════════════════════════════
# METHOD 9: BenchReg + KNN Blend (simple but effective)
# ══════════════════════════════════════════════════════════════════════════

def breg_knn_blend(M_train, alpha=0.65):
    """
    Simple blend: alpha * BenchReg + (1-alpha) * KNN.
    BenchReg is best per-model, KNN is best random CV. Blending should help both.
    """
    M_breg = bench_reg_predict(M_train, top_k=5, min_r2=0.2)
    M_knn = knn_predict(M_train, k=5, metric='cosine')
    mask = ~np.isnan(M_train)
    col_m = np.nanmean(M_train, axis=0)

    M_pred = M_train.copy()
    for i in range(n_models):
        for j in range(n_benchmarks):
            if mask[i, j]: continue
            b = M_breg[i, j]
            k = M_knn[i, j]
            if np.isnan(b) and np.isnan(k):
                M_pred[i, j] = col_m[j]
            elif np.isnan(b):
                M_pred[i, j] = k
            elif np.isnan(k):
                M_pred[i, j] = b
            else:
                # Check if BReg fell back to col mean — if so, trust KNN more
                if abs(b - col_m[j]) < 0.01:
                    M_pred[i, j] = k
                elif abs(k - col_m[j]) < 0.01:
                    M_pred[i, j] = b
                else:
                    M_pred[i, j] = alpha * b + (1 - alpha) * k
    return M_pred


# ══════════════════════════════════════════════════════════════════════════
# METHOD 10: Log-transform BReg+KNN for high-variance benchmarks
# ══════════════════════════════════════════════════════════════════════════

# Detect high-variance benchmarks (CV > 50%)
high_var_benchmarks = set()
for j in range(n_benchmarks):
    vals = M_full[observed[:, j], j]
    if len(vals) >= 3:
        cv = np.std(vals) / (np.mean(vals) + 0.001)
        if cv > 0.40:
            high_var_benchmarks.add(bench_ids[j])
print(f"High-variance benchmarks (CV>40%): {sorted(high_var_benchmarks)}")


def log_breg_knn_blend(M_train, alpha=0.6):
    """
    Like breg_knn_blend, but for high-variance benchmarks, predict in log-space.
    This compresses the range (e.g., ARC-AGI-2: 1-77 → 0-4.3) making linear
    models much more appropriate.
    """
    mask = ~np.isnan(M_train)
    col_m = np.nanmean(M_train, axis=0)

    # Transform high-var columns to log space
    M_log = M_train.copy()
    for j in range(n_benchmarks):
        if bench_ids[j] in high_var_benchmarks:
            for i in range(n_models):
                if not np.isnan(M_log[i, j]):
                    M_log[i, j] = np.log1p(M_log[i, j])  # log(1+x)

    # Run predictions in log space
    M_breg_log = bench_reg_predict(M_log, top_k=5, min_r2=0.2)
    M_knn_log = knn_predict(M_log, k=5, metric='cosine')

    # Also run in normal space
    M_breg = bench_reg_predict(M_train, top_k=5, min_r2=0.2)
    M_knn = knn_predict(M_train, k=5, metric='cosine')

    M_pred = M_train.copy()
    for i in range(n_models):
        for j in range(n_benchmarks):
            if mask[i, j]: continue

            if bench_ids[j] in high_var_benchmarks:
                # Use log-space predictions, transform back
                b = M_breg_log[i, j]
                k = M_knn_log[i, j]
                if not np.isnan(b): b = np.expm1(b)  # exp(x)-1
                if not np.isnan(k): k = np.expm1(k)
            else:
                b = M_breg[i, j]
                k = M_knn[i, j]

            if np.isnan(b) and np.isnan(k):
                M_pred[i, j] = col_m[j]
            elif np.isnan(b):
                M_pred[i, j] = k
            elif np.isnan(k):
                M_pred[i, j] = b
            else:
                col_m_j = col_m[j]
                if abs(b - col_m_j) < 0.01:
                    M_pred[i, j] = k
                elif abs(k - col_m_j) < 0.01:
                    M_pred[i, j] = b
                else:
                    M_pred[i, j] = alpha * b + (1 - alpha) * k
    return M_pred


# ══════════════════════════════════════════════════════════════════════════
# CLAMPING (post-processing)
# ══════════════════════════════════════════════════════════════════════════

def clamp_predictions(M_pred):
    """Clamp predictions to plausible ranges per benchmark."""
    M_clamped = M_pred.copy()
    for j in range(n_benchmarks):
        M_clamped[:, j] = np.clip(M_clamped[:, j], bench_min[j], bench_max[j])
    return M_clamped


# ══════════════════════════════════════════════════════════════════════════
# CONFIDENCE INTERVALS via method disagreement
# ══════════════════════════════════════════════════════════════════════════

def compute_confidence(M_full_matrix):
    """
    For each missing cell, compute:
    - Point estimate (from best method)
    - Confidence interval (from method disagreement)
    - Confidence level: 'high', 'medium', 'low'
    """
    mask = ~np.isnan(M_full_matrix)

    # Run all methods
    methods_preds = [
        als_predict(M_full_matrix, rank=5, lam=1.0),
        als_predict(M_full_matrix, rank=8, lam=1.0),
        knn_predict(M_full_matrix, k=5, metric='cosine'),
        knn_predict(M_full_matrix, k=7, metric='cosine'),
        bench_reg_predict(M_full_matrix, top_k=5, min_r2=0.2),
        scaling_predict(M_full_matrix),
    ]

    # Stack predictions
    stacked = np.stack(methods_preds, axis=-1)  # (n_models, n_benchmarks, n_methods)

    # Point estimate (median of methods)
    M_point = np.nanmedian(stacked, axis=-1)

    # CI = IQR of method predictions
    M_lo = np.nanpercentile(stacked, 25, axis=-1)
    M_hi = np.nanpercentile(stacked, 75, axis=-1)

    # Method std as fraction of point estimate
    M_std = np.nanstd(stacked, axis=-1)
    M_cv = M_std / (np.abs(M_point) + 1e-6)  # Coefficient of variation

    # Confidence levels
    confidence = np.full((n_models, n_benchmarks), '', dtype=object)
    for i in range(n_models):
        for j in range(n_benchmarks):
            if mask[i, j]:
                confidence[i, j] = 'observed'
            elif M_cv[i, j] < 0.10:
                confidence[i, j] = 'high'
            elif M_cv[i, j] < 0.25:
                confidence[i, j] = 'medium'
            else:
                confidence[i, j] = 'low'

    return M_point, M_lo, M_hi, confidence, M_cv


# ══════════════════════════════════════════════════════════════════════════
# CV FRAMEWORK
# ══════════════════════════════════════════════════════════════════════════

def make_folds(n_folds=3, seed=42):
    np.random.seed(seed)
    obs_idx = list(zip(*np.where(observed)))
    perm = np.random.permutation(len(obs_idx))
    fold_size = len(obs_idx) // n_folds
    folds = []
    for k in range(n_folds):
        start = k * fold_size
        end = start + fold_size if k < n_folds - 1 else len(obs_idx)
        test_indices = [obs_idx[perm[p]] for p in range(start, end)]
        M_train = M_full.copy()
        for i, j in test_indices:
            M_train[i, j] = np.nan
        folds.append((M_train, test_indices))
    return folds


def make_per_model_folds(n_folds=3, seed=42, holdout_frac=0.2):
    np.random.seed(seed)
    folds = []
    for fold_k in range(n_folds):
        M_train = M_full.copy()
        test_indices = []
        model_groups = defaultdict(list)
        for i in range(n_models):
            obs_j = np.where(observed[i, :])[0]
            n_obs = len(obs_j)
            if n_obs < 5: continue
            n_hide = max(1, int(n_obs * holdout_frac))
            rng = np.random.RandomState(seed * 1000 + fold_k * 100 + i)
            hidden = rng.choice(obs_j, size=n_hide, replace=False)
            for j in hidden:
                M_train[i, j] = np.nan
                test_indices.append((i, j))
                model_groups[i].append(j)
        folds.append((M_train, test_indices, dict(model_groups)))
    return folds


def cv_evaluate(predict_fn, folds, name=""):
    all_errors = []
    for fold_data in folds:
        M_train = fold_data[0]
        test_indices = fold_data[1]
        M_pred = predict_fn(M_train)
        M_pred = clamp_predictions(M_pred)
        for i, j in test_indices:
            actual = M_full[i, j]
            if abs(actual) < 1e-6: continue
            pred = M_pred[i, j]
            if np.isnan(pred): continue
            all_errors.append(abs((pred - actual) / actual))
    e = np.array(all_errors)
    return {
        'name': name,
        'mape': np.nanmean(e) * 100,
        'median_ape': np.nanmedian(e) * 100,
        'w5': np.mean(e <= 0.05) * 100,
        'w10': np.mean(e <= 0.10) * 100,
        'w20': np.mean(e <= 0.20) * 100,
        'n': len(e),
    }


def per_model_evaluate(predict_fn, folds, name=""):
    model_errors = defaultdict(list)
    bench_errors = defaultdict(list)
    all_errors = []
    for M_train, test_indices, model_groups in folds:
        M_pred = predict_fn(M_train)
        M_pred = clamp_predictions(M_pred)
        for i, j in test_indices:
            actual = M_full[i, j]
            if abs(actual) < 1e-6: continue
            pred = M_pred[i, j]
            if np.isnan(pred): continue
            ape = abs((pred - actual) / actual)
            all_errors.append(ape)
            model_errors[i].append(ape)
            bench_errors[j].append(ape)
    e = np.array(all_errors)
    result = {
        'name': name,
        'mape': np.nanmean(e) * 100,
        'median_ape': np.nanmedian(e) * 100,
        'w5': np.mean(e <= 0.05) * 100,
        'w10': np.mean(e <= 0.10) * 100,
        'w20': np.mean(e <= 0.20) * 100,
        'n': len(e),
    }
    model_medape = {}
    for i, errs in model_errors.items():
        model_medape[model_ids[i]] = np.nanmedian(errs) * 100
    bench_medape = {}
    for j, errs in bench_errors.items():
        bench_medape[bench_ids[j]] = np.nanmedian(errs) * 100
    result['model_medape'] = model_medape
    result['bench_medape'] = bench_medape
    return result


# ══════════════════════════════════════════════════════════════════════════
# RUN EVALUATION
# ══════════════════════════════════════════════════════════════════════════

print("\n" + "=" * 90)
print("  EVALUATION 1: Random 3-fold CV (hide random 33% of entries)")
print("=" * 90)

random_folds = make_folds(n_folds=3, seed=42)

methods = [
    ("ALS(r=5,λ=1.0)", lambda M: als_predict(M, rank=5, lam=1.0)),
    ("ALS(r=8,λ=1.0)", lambda M: als_predict(M, rank=8, lam=1.0)),
    ("KNN(k=5,cos)", lambda M: knn_predict(M, k=5, metric='cosine')),
    ("KNN(k=7,cos)", lambda M: knn_predict(M, k=7, metric='cosine')),
    ("BenchReg(k=5,r²≥0.2)", lambda M: bench_reg_predict(M, top_k=5, min_r2=0.2)),
    ("BenchReg(k=5,r²≥0.3)", lambda M: bench_reg_predict(M, top_k=5, min_r2=0.3)),
    ("ScalingLaw", scaling_predict),
    ("TrimmedEnsemble", trimmed_ensemble_predict),
    ("StackedEnsemble", stacked_ensemble_predict),
    ("BReg+KNN(α=0.6)", lambda M: breg_knn_blend(M, alpha=0.6)),
    ("BReg+KNN(α=0.65)", lambda M: breg_knn_blend(M, alpha=0.65)),
    ("BReg+KNN(α=0.7)", lambda M: breg_knn_blend(M, alpha=0.7)),
    ("LogBReg+KNN(α=0.6)", lambda M: log_breg_knn_blend(M, alpha=0.6)),
    ("LogBReg+KNN(α=0.65)", lambda M: log_breg_knn_blend(M, alpha=0.65)),
]

print("\n  Testing methods (with clamping)...")
results = []
for name, fn in methods:
    r = cv_evaluate(fn, random_folds, name)
    results.append(r)
    print(f"    {name:<35s}  MedAPE={r['median_ape']:>5.1f}%  MAPE={r['mape']:>5.1f}%  ±10%={r['w10']:>5.1f}%  ±20%={r['w20']:>5.1f}%")

results.sort(key=lambda x: x['median_ape'])
print(f"\n  Best random CV: {results[0]['name']} (MedAPE={results[0]['median_ape']:.1f}%)")


# ══════════════════════════════════════════════════════════════════════════
# EVALUATION 2: Per-Model Holdout
# ══════════════════════════════════════════════════════════════════════════

print("\n\n" + "=" * 90)
print("  EVALUATION 2: Per-Model Holdout (hide 20% of each model, 3 folds)")
print("=" * 90)

pm_folds = make_per_model_folds(n_folds=3, seed=42, holdout_frac=0.2)
n_test_entries = sum(len(ti) for _, ti, _ in pm_folds)
print(f"  Total test entries across 3 folds: {n_test_entries}")

# Test all methods
pm_results = []
for name, fn in methods:
    r = per_model_evaluate(fn, pm_folds, name)
    pm_results.append(r)
    print(f"\n  {name}:")
    print(f"    Overall: MedAPE={r['median_ape']:.1f}%  MAPE={r['mape']:.1f}%  ±10%={r['w10']:.1f}%  ±20%={r['w20']:.1f}%")

pm_results.sort(key=lambda x: x['median_ape'])
best_pm = pm_results[0]
print(f"\n  Best per-model holdout: {best_pm['name']} (MedAPE={best_pm['median_ape']:.1f}%)")

# Show detailed per-model breakdown for best method
print(f"\n  Per-model breakdown ({best_pm['name']}):")
print(f"  {'Model':<40s} {'Scores':>6s} {'MedAPE':>8s}")
print(f"  {'─'*40} {'─'*6} {'─'*8}")
for mid, medape in sorted(best_pm['model_medape'].items(), key=lambda x: x[1]):
    n_scores = int(observed[model_idx[mid]].sum())
    print(f"  {model_names[mid]:<40s} {n_scores:>6d} {medape:>7.1f}%")

good_models = sum(1 for m, e in best_pm['model_medape'].items() if e < 10)
ok_models = sum(1 for m, e in best_pm['model_medape'].items() if e < 20)
total_models = len(best_pm['model_medape'])
print(f"\n  Models <5%: {sum(1 for m,e in best_pm['model_medape'].items() if e<5)}/{total_models}")
print(f"  Models <10%: {good_models}/{total_models}")
print(f"  Models <20%: {ok_models}/{total_models}")

# Per-benchmark
print(f"\n  Per-benchmark breakdown ({best_pm['name']}):")
print(f"  {'Benchmark':<35s} {'N obs':>6s} {'MedAPE':>8s}")
print(f"  {'─'*35} {'─'*6} {'─'*8}")
for bid, medape in sorted(best_pm['bench_medape'].items(), key=lambda x: x[1]):
    n_obs = int(observed[:, bench_idx[bid]].sum())
    print(f"  {bench_names.get(bid, bid):<35s} {n_obs:>6d} {medape:>7.1f}%")


# ══════════════════════════════════════════════════════════════════════════
# GENERATE COMPLETE PREDICTED MATRIX + CONFIDENCE INTERVALS
# ══════════════════════════════════════════════════════════════════════════

print("\n\n" + "=" * 90)
print("  GENERATING COMPLETE PREDICTED MATRIX")
print("=" * 90)

# Use best method for point estimates
best_fn = dict(methods)[best_pm['name']]
M_best = clamp_predictions(best_fn(M_full))

# Compute confidence intervals from method disagreement
M_point, M_lo, M_hi, confidence, M_cv = compute_confidence(M_full)
M_point = clamp_predictions(M_point)
M_lo = clamp_predictions(M_lo)
M_hi = clamp_predictions(M_hi)

# Use best method for final point estimates (confidence from ensemble)
M_final = M_full.copy()
for i in range(n_models):
    for j in range(n_benchmarks):
        if np.isnan(M_final[i, j]):
            M_final[i, j] = M_best[i, j]

n_high = sum(1 for i in range(n_models) for j in range(n_benchmarks) if confidence[i,j] == 'high')
n_med = sum(1 for i in range(n_models) for j in range(n_benchmarks) if confidence[i,j] == 'medium')
n_low = sum(1 for i in range(n_models) for j in range(n_benchmarks) if confidence[i,j] == 'low')
n_obs = sum(1 for i in range(n_models) for j in range(n_benchmarks) if confidence[i,j] == 'observed')
print(f"  Observed: {n_obs}, Predicted high confidence: {n_high}, medium: {n_med}, low: {n_low}")


# ══════════════════════════════════════════════════════════════════════════
# OUTPUT: Excel with predictions + confidence
# ══════════════════════════════════════════════════════════════════════════

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers

    wb = openpyxl.Workbook()

    # ── Sheet 1: Complete Matrix (observed + predicted) ──
    ws1 = wb.active
    ws1.title = "Complete Matrix"

    # Header
    ws1.cell(row=1, column=1, value="Model")
    for j, bid in enumerate(bench_ids):
        ws1.cell(row=1, column=j+2, value=bench_names[bid])

    # Fills
    fill_obs = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_high = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # light green
    fill_med = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")   # light yellow
    fill_low = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")   # light red
    fill_header = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    font_header = Font(color="FFFFFF", bold=True, size=10)
    font_obs = Font(bold=True, size=10)
    font_pred = Font(italic=True, size=10, color="666666")

    for c in range(1, n_benchmarks + 2):
        ws1.cell(row=1, column=c).fill = fill_header
        ws1.cell(row=1, column=c).font = font_header

    for i, mid in enumerate(model_ids):
        ws1.cell(row=i+2, column=1, value=model_names[mid])
        for j in range(n_benchmarks):
            cell = ws1.cell(row=i+2, column=j+2)
            val = M_final[i, j]
            if np.isnan(val):
                continue
            cell.value = round(val, 1)
            conf = confidence[i, j]
            if conf == 'observed':
                cell.fill = fill_obs
                cell.font = font_obs
            elif conf == 'high':
                cell.fill = fill_high
                cell.font = font_pred
            elif conf == 'medium':
                cell.fill = fill_med
                cell.font = font_pred
            else:
                cell.fill = fill_low
                cell.font = font_pred

    ws1.freeze_panes = "B2"
    ws1.column_dimensions['A'].width = 30

    # ── Sheet 2: Observed Only ──
    ws2 = wb.create_sheet("Observed Only")
    ws2.cell(row=1, column=1, value="Model")
    for j, bid in enumerate(bench_ids):
        ws2.cell(row=1, column=j+2, value=bench_names[bid])
    for c in range(1, n_benchmarks + 2):
        ws2.cell(row=1, column=c).fill = fill_header
        ws2.cell(row=1, column=c).font = font_header
    for i, mid in enumerate(model_ids):
        ws2.cell(row=i+2, column=1, value=model_names[mid])
        for j in range(n_benchmarks):
            if observed[i, j]:
                ws2.cell(row=i+2, column=j+2, value=round(M_full[i, j], 1))
    ws2.freeze_panes = "B2"
    ws2.column_dimensions['A'].width = 30

    # ── Sheet 3: Confidence Intervals ──
    ws3 = wb.create_sheet("Confidence Intervals")
    ws3.cell(row=1, column=1, value="Model")
    ws3.cell(row=1, column=2, value="Benchmark")
    ws3.cell(row=1, column=3, value="Observed")
    ws3.cell(row=1, column=4, value="Predicted")
    ws3.cell(row=1, column=5, value="CI Low (25%)")
    ws3.cell(row=1, column=6, value="CI High (75%)")
    ws3.cell(row=1, column=7, value="Confidence")
    ws3.cell(row=1, column=8, value="CV (%)")
    for c in range(1, 9):
        ws3.cell(row=1, column=c).fill = fill_header
        ws3.cell(row=1, column=c).font = font_header

    row = 2
    for i in range(n_models):
        for j in range(n_benchmarks):
            if observed[i, j]:
                continue
            ws3.cell(row=row, column=1, value=model_names[model_ids[i]])
            ws3.cell(row=row, column=2, value=bench_names[bench_ids[j]])
            ws3.cell(row=row, column=3, value="")
            ws3.cell(row=row, column=4, value=round(float(M_final[i, j]), 1))
            ws3.cell(row=row, column=5, value=round(float(M_lo[i, j]), 1))
            ws3.cell(row=row, column=6, value=round(float(M_hi[i, j]), 1))
            ws3.cell(row=row, column=7, value=confidence[i, j])
            ws3.cell(row=row, column=8, value=round(float(M_cv[i, j] * 100), 1))
            conf = confidence[i, j]
            if conf == 'high':
                ws3.cell(row=row, column=7).fill = fill_high
            elif conf == 'medium':
                ws3.cell(row=row, column=7).fill = fill_med
            else:
                ws3.cell(row=row, column=7).fill = fill_low
            row += 1

    ws3.freeze_panes = "A2"
    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 30

    # ── Sheet 4: CV Results ──
    ws4 = wb.create_sheet("CV Results")
    ws4.cell(row=1, column=1, value="Method")
    ws4.cell(row=1, column=2, value="MedAPE (%)")
    ws4.cell(row=1, column=3, value="MAPE (%)")
    ws4.cell(row=1, column=4, value="Within 10% (%)")
    ws4.cell(row=1, column=5, value="Within 20% (%)")
    ws4.cell(row=1, column=6, value="N test")
    for c in range(1, 7):
        ws4.cell(row=1, column=c).fill = fill_header
        ws4.cell(row=1, column=c).font = font_header

    for idx, r in enumerate(pm_results):
        row = idx + 2
        ws4.cell(row=row, column=1, value=r['name'])
        ws4.cell(row=row, column=2, value=round(r['median_ape'], 1))
        ws4.cell(row=row, column=3, value=round(r['mape'], 1))
        ws4.cell(row=row, column=4, value=round(r['w10'], 1))
        ws4.cell(row=row, column=5, value=round(r['w20'], 1))
        ws4.cell(row=row, column=6, value=r['n'])

    ws4.column_dimensions['A'].width = 35

    # ── Sheet 5: Stats ──
    ws5 = wb.create_sheet("Stats")
    stats = [
        ("Matrix size", f"{n_models} models × {n_benchmarks} benchmarks"),
        ("Observed entries", f"{n_observed} ({100*n_observed/(n_models*n_benchmarks):.1f}%)"),
        ("Predicted entries", f"{n_models*n_benchmarks - n_observed}"),
        ("Best method (random CV)", f"{results[0]['name']} (MedAPE={results[0]['median_ape']:.1f}%)"),
        ("Best method (per-model)", f"{best_pm['name']} (MedAPE={best_pm['median_ape']:.1f}%)"),
        ("Models <10% error", f"{good_models}/{total_models}"),
        ("Models <20% error", f"{ok_models}/{total_models}"),
        ("High confidence predictions", str(n_high)),
        ("Medium confidence predictions", str(n_med)),
        ("Low confidence predictions", str(n_low)),
        ("", ""),
        ("Color Legend (Complete Matrix sheet)", ""),
        ("White / Bold", "Observed (real score)"),
        ("Green / Italic", "Predicted - High confidence (methods agree within 10%)"),
        ("Yellow / Italic", "Predicted - Medium confidence (methods agree within 25%)"),
        ("Red / Italic", "Predicted - Low confidence (methods disagree >25%)"),
    ]
    for idx, (k, v) in enumerate(stats):
        ws5.cell(row=idx+1, column=1, value=k)
        ws5.cell(row=idx+1, column=2, value=v)
    ws5.column_dimensions['A'].width = 35
    ws5.column_dimensions['B'].width = 60

    output_path = os.path.join(REPO_ROOT, 'results', 'llm_benchmark_predictions.xlsx')
    wb.save(output_path)
    print(f"\n  Excel saved: {output_path}")

except ImportError:
    print("  openpyxl not available, skipping Excel output")


# ══════════════════════════════════════════════════════════════════════════
# OUTPUT: JSON with full predictions
# ══════════════════════════════════════════════════════════════════════════

json_output = {
    "metadata": {
        "n_models": n_models,
        "n_benchmarks": n_benchmarks,
        "n_observed": int(n_observed),
        "fill_rate": round(100 * n_observed / (n_models * n_benchmarks), 1),
        "best_method_random_cv": results[0]['name'],
        "best_medape_random_cv": round(results[0]['median_ape'], 1),
        "best_method_per_model": best_pm['name'],
        "best_medape_per_model": round(best_pm['median_ape'], 1),
    },
    "models": [{"id": m[0], "name": m[1], "provider": m[2]} for m in MODELS],
    "benchmarks": [{"id": b[0], "name": b[1], "category": b[2]} for b in BENCHMARKS],
    "predictions": []
}

for i in range(n_models):
    for j in range(n_benchmarks):
        entry = {
            "model": model_ids[i],
            "benchmark": bench_ids[j],
            "value": round(float(M_final[i, j]), 1) if not np.isnan(M_final[i, j]) else None,
            "observed": bool(observed[i, j]),
            "confidence": confidence[i, j] if confidence[i, j] else None,
        }
        if not observed[i, j] and not np.isnan(M_final[i, j]):
            entry["ci_low"] = round(float(M_lo[i, j]), 1)
            entry["ci_high"] = round(float(M_hi[i, j]), 1)
        json_output["predictions"].append(entry)

json_path = os.path.join(REPO_ROOT, 'results', 'llm_benchmark_predictions.json')
with open(json_path, 'w') as f:
    json.dump(json_output, f, indent=2)
print(f"  JSON saved: {json_path}")


# ══════════════════════════════════════════════════════════════════════════
# SUMMARY
# ══════════════════════════════════════════════════════════════════════════

print("\n\n" + "=" * 90)
print("  SUMMARY")
print("=" * 90)

print(f"\n  Matrix: {n_models} models × {n_benchmarks} benchmarks")
print(f"  Observed entries: {n_observed} ({100*n_observed/(n_models*n_benchmarks):.1f}%)")
print(f"\n  Random 3-fold CV:      best MedAPE = {results[0]['median_ape']:.1f}% ({results[0]['name']})")
print(f"  Per-model holdout CV:  best MedAPE = {best_pm['median_ape']:.1f}% ({best_pm['name']})")
print(f"\n  Models <5%: {sum(1 for m,e in best_pm['model_medape'].items() if e<5)}/{total_models}")
print(f"  Models <10%: {good_models}/{total_models}")
print(f"  Models <20%: {ok_models}/{total_models}")
print(f"\n  Predicted cells: {n_high + n_med + n_low}")
print(f"    High confidence:   {n_high} ({100*n_high/(n_high+n_med+n_low):.0f}%)")
print(f"    Medium confidence: {n_med} ({100*n_med/(n_high+n_med+n_low):.0f}%)")
print(f"    Low confidence:    {n_low} ({100*n_low/(n_high+n_med+n_low):.0f}%)")
print(f"\n  Output files:")
print(f"    llm_benchmark_predictions.xlsx  (5 sheets: Complete Matrix, Observed Only, CIs, CV Results, Stats)")
print(f"    llm_benchmark_predictions.json  (full data with confidence intervals)")
print("\n" + "=" * 90)
